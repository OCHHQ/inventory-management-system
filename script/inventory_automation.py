"""
INVENTORY AUTOMATION SYSTEM
============================
A complete Python automation for inventory tracking and management.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime
import os

class InventoryManager:
    """Main class for inventory management operations"""
    
    def __init__(self, excel_file):
        self.excel_file = excel_file
        self.wb = load_workbook(excel_file)
    
    def add_item(self, sheet_name, item_name, quantity, unit, unit_price=0, reorder_level=5):
        """Add a new item to specified sheet"""
        sheet = self.wb[sheet_name]
        
        # Find next empty row
        next_row = sheet.max_row + 1
        
        # Check if we need to insert before summary section
        for row in range(sheet.max_row, 5, -1):
            if sheet.cell(row=row, column=4).value == 'TOTAL:':
                next_row = row
                break
        
        # Add item data
        sheet.cell(row=next_row, column=1, value=item_name)
        sheet.cell(row=next_row, column=2, value=quantity)
        sheet.cell(row=next_row, column=3, value=unit)
        
        # Unit price (blue for input)
        price_cell = sheet.cell(row=next_row, column=4, value=unit_price)
        price_cell.font = Font(color='0000FF')
        
        # Total formula
        total_cell = sheet.cell(row=next_row, column=5, value=f'=B{next_row}*D{next_row}')
        total_cell.font = Font(color='000000')
        total_cell.number_format = '₦#,##0;(₦#,##0);-'
        
        # Reorder level
        reorder_cell = sheet.cell(row=next_row, column=6, value=reorder_level)
        reorder_cell.font = Font(color='0000FF')
        
        # Status formula
        status_cell = sheet.cell(row=next_row, column=7, value=f'=IF(B{next_row}<F{next_row},"REORDER","OK")')
        status_cell.font = Font(color='000000')
        
        print(f"✅ Added {item_name} to {sheet_name}")
    
    def update_quantity(self, sheet_name, item_name, new_quantity):
        """Update quantity for an existing item"""
        sheet = self.wb[sheet_name]
        
        # Find item
        for row in range(6, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == item_name:
                sheet.cell(row=row, column=2, value=new_quantity)
                print(f"✅ Updated {item_name} quantity to {new_quantity}")
                return
        
        print(f"❌ Item '{item_name}' not found in {sheet_name}")
    
    def update_price(self, sheet_name, item_name, new_price):
        """Update unit price for an existing item"""
        sheet = self.wb[sheet_name]
        
        for row in range(6, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == item_name:
                price_cell = sheet.cell(row=row, column=4, value=new_price)
                price_cell.font = Font(color='0000FF')
                print(f"✅ Updated {item_name} price to ₦{new_price}")
                return
        
        print(f"❌ Item '{item_name}' not found in {sheet_name}")
    
    def get_items_to_reorder(self, sheet_name=None):
        """Get list of items that need reordering"""
        sheets_to_check = [sheet_name] if sheet_name else ['Cashier Items', 'Cleaning Items']
        reorder_items = []
        
        for sname in sheets_to_check:
            sheet = self.wb[sname]
            
            for row in range(6, sheet.max_row + 1):
                item = sheet.cell(row=row, column=1).value
                qty = sheet.cell(row=row, column=2).value
                reorder_lvl = sheet.cell(row=row, column=6).value
                
                if item and qty and reorder_lvl:
                    # Handle string quantities like "12 & 3"
                    try:
                        qty_num = float(str(qty).split()[0]) if isinstance(qty, str) else float(qty)
                        if qty_num < reorder_lvl:
                            reorder_items.append({
                                'Category': sname,
                                'Item': item,
                                'Current Qty': qty,
                                'Reorder Level': reorder_lvl
                            })
                    except:
                        pass
        
        return pd.DataFrame(reorder_items)
    
    def generate_report(self, output_file='inventory_report.txt'):
        """Generate text report of current inventory status"""
        report_lines = []
        report_lines.append("=" * 60)
        report_lines.append("INVENTORY STATUS REPORT")
        report_lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        report_lines.append("=" * 60)
        
        for sheet_name in ['Cashier Items', 'Cleaning Items']:
            sheet = self.wb[sheet_name]
            report_lines.append(f"\n{sheet_name.upper()}")
            report_lines.append("-" * 60)
            
            for row in range(6, sheet.max_row + 1):
                item = sheet.cell(row=row, column=1).value
                qty = sheet.cell(row=row, column=2).value
                unit = sheet.cell(row=row, column=3).value
                
                if item and item != 'TOTAL:':
                    report_lines.append(f"  {item:.<40} {qty} {unit}")
        
        # Items to reorder
        reorder_df = self.get_items_to_reorder()
        if not reorder_df.empty:
            report_lines.append("\n" + "=" * 60)
            report_lines.append("⚠️  ITEMS REQUIRING REORDER")
            report_lines.append("=" * 60)
            for _, row in reorder_df.iterrows():
                report_lines.append(f"  [{row['Category']}] {row['Item']}: {row['Current Qty']} (Min: {row['Reorder Level']})")
        
        report = "\n".join(report_lines)
        
        with open(output_file, 'w') as f:
            f.write(report)
        
        print(f"📄 Report saved to {output_file}")
        return report
    
    def export_to_csv(self, output_dir='exports'):
        """Export each sheet to CSV"""
        os.makedirs(output_dir, exist_ok=True)
        
        for sheet_name in ['Cashier Items', 'Cleaning Items']:
            df = pd.read_excel(self.excel_file, sheet_name=sheet_name, skiprows=4)
            csv_file = f"{output_dir}/{sheet_name.lower().replace(' ', '_')}.csv"
            df.to_csv(csv_file, index=False)
            print(f"💾 Exported {sheet_name} to {csv_file}")
    
    def save(self):
        """Save changes to Excel file"""
        self.wb.save(self.excel_file)
        print(f"💾 Saved changes to {self.excel_file}")


# ===========================
# USAGE EXAMPLES
# ===========================

if __name__ == "__main__":
    # Initialize the manager
    manager = InventoryManager('inventory_tracker.xlsx')
    
    print("\n" + "="*60)
    print("INVENTORY AUTOMATION SYSTEM - DEMO")
    print("="*60)
    
    # Example 1: Add new items
    print("\n📦 Adding new items...")
    manager.add_item('Cashier Items', 'Highlighter', 12, 'pic', unit_price=150, reorder_level=5)
    manager.add_item('Cleaning Items', 'Jik', 7,'bottles', unit_price=2600, reorder_level=5)
    manager.add_item('Cleaning Items', 'Big HandTissue', 89, 'pieces', unit_price=2000, reorder_level=40 )
    manager.add_item('Cleaning Items', 'Tissue Toilet paper', 24, 'pieces', unit_price=1500, reorder_level=40)
    manager.add_item('Cleaning Items', 'Bucket', 7, 'pieces', unit_price=1500, reorder_level=6)
    manager.add_item('Cleaning Items', 'Dust bin Leather', 20, 'pieces', unit_price= 700, reorder_level=20)
    
    # Example 2: Update quantities
    print("\n📊 Updating quantities...")
    manager.update_quantity('Cashier Items', 'Marker', 15)
    manager.update_quantity('Cleaning Items', 'Coffee', 40)
    
    
    # Example 3: Update prices
    print("\n💰 Updating prices...")
    manager.update_price('Cashier Items', 'A4 Paper', 2500)
    manager.update_price('Cleaning Items', 'Vim', 450)
    
    # Example 4: Get items to reorder
    print("\n⚠️  Checking items that need reordering...")
    reorder_df = manager.get_items_to_reorder()
    if not reorder_df.empty:
        print(reorder_df.to_string(index=False))
    else:
        print("✅ All items are adequately stocked!")
    
    # Example 5: Generate report
    print("\n📄 Generating inventory report...")
    report = manager.generate_report()
    print("\nReport preview:")
    print(report[:500] + "...")
    
    # Example 6: Export to CSV
    print("\n💾 Exporting to CSV...")
    manager.export_to_csv()
    
    # Save all changes
    print("\n💾 Saving changes...")
    manager.save()
    
    print("\n✅ Demo completed successfully!")
    print("="*60)
