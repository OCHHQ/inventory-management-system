from openpyxl import load_workbook
from openpyxl.styles import Font

print("🔧 Fixing Cleaning Items summary formulas...\n")

# Load workbook
wb = load_workbook('inventory_tracker.xlsx')
cleaning = wb['Cleaning Items']

# Last item is at row 28 (Dust bin Leather)
last_item_row = 28

# Summary should be at row 30-31
summary_row = 30
reorder_row = 31

print(f"📊 Items in rows 6 to {last_item_row}")
print(f"📊 Moving summary to rows {summary_row} and {reorder_row}\n")

# Clear row 29 (the wrong "Items to Reorder" location)
for col in range(1, 8):
    cleaning.cell(row=29, column=col, value=None)

# Add TOTAL row at row 30
cleaning.cell(row=summary_row, column=4, value='TOTAL:').font = Font(bold=True)
total_cell = cleaning.cell(row=summary_row, column=5)
total_cell.value = f'=SUM(E6:E{last_item_row})'
total_cell.font = Font(bold=True, color='000000')
total_cell.number_format = '₦#,##0;(₦#,##0);-'
print(f"✅ Row {summary_row}: TOTAL = =SUM(E6:E{last_item_row})")

# Add Items to Reorder row at row 31
cleaning.cell(row=reorder_row, column=4, value='Items to Reorder:').font = Font(bold=True)
reorder_cell = cleaning.cell(row=reorder_row, column=5)
reorder_cell.value = f'=COUNTIF(G6:G{last_item_row},"REORDER")'
reorder_cell.font = Font(bold=True, color='000000')
print(f"✅ Row {reorder_row}: Items to Reorder = =COUNTIF(G6:G{last_item_row},\"REORDER\")")

# Fix Dashboard references
dashboard = wb['Dashboard']

print("\n🎯 Updating Dashboard references...")

# Cleaning Total Cost (B10)
dashboard['B10'].value = f"='Cleaning Items'!E{summary_row}"
dashboard['B10'].number_format = '₦#,##0;(₦#,##0);-'
print(f"✅ Dashboard B10 = 'Cleaning Items'!E{summary_row}")

# Cleaning Items to Reorder (B11)
dashboard['B11'].value = f"='Cleaning Items'!E{reorder_row}"
print(f"✅ Dashboard B11 = 'Cleaning Items'!E{reorder_row}")

# Grand Total (B13)
dashboard['B13'].value = '=B6+B10'
dashboard['B13'].font = Font(bold=True, size=12, color='FF0000')
dashboard['B13'].number_format = '₦#,##0;(₦#,##0);-'
print(f"✅ Dashboard B13 = =B6+B10")

# Save
wb.save('inventory_tracker.xlsx')

print("\n" + "=" * 60)
print("✅ ALL FIXED!")
print("=" * 60)
print("\nSummary:")
print(f"  • Cleaning Items: {last_item_row - 5} items")
print(f"  • TOTAL formula at row {summary_row}")
print(f"  • Items to Reorder at row {reorder_row}")
print(f"  • Dashboard updated")
print("\n💾 File saved: inventory_tracker.xlsx\n")
